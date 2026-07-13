import calendar
import io
import json
from datetime import date, datetime, time, timedelta

from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.db import transaction
from django.db.models.functions import Lower
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.utils.decorators import method_decorator
from django.views import View

from .models import Child, Therapist, Therapy


def _slot_hours():
    """All hours from 08:00 to 20:00."""
    return [f"{h:02d}:00" for h in range(8, 21)]


def _therapies_for_date(session_date: date) -> list:
    """Return existing therapies for a date grouped by therapist, ordered by start_time."""
    rows = (
        Therapy.objects.filter(date=session_date)
        .select_related("therapist", "child")
        .order_by("therapist__last_name", "therapist__first_name", "start_time")
    )
    groups = {}
    for t in rows:
        tid = t.therapist_id
        if tid not in groups:
            groups[tid] = {"therapist_id": tid, "slots": []}
        groups[tid]["slots"].append({
            "child_id": t.child_id,
            "start_time": t.start_time.strftime("%H:%M"),
        })
    return list(groups.values())


@method_decorator(staff_member_required, name="dispatch")
class TherapyBatchView(View):
    template_name = "admin/core/therapy_batch.html"

    def _base_context(self, request):
        return {
            **admin_site_context(request),
            "therapists": list(
                Therapist.objects.order_by(Lower("last_name"), Lower("first_name")).values("id", "first_name", "last_name")
            ),
            "children": list(
                Child.objects.filter(status="activ")
                .order_by(Lower("last_name"), Lower("first_name"))
                .values("id", "first_name", "last_name")
            ),
            "slot_hours": json.dumps(_slot_hours()),
            "today": date.today().isoformat(),
            "tomorrow": (date.today() + timedelta(days=1)).isoformat(),
        }

    def get(self, request):
        # AJAX: return existing therapies for a given date as JSON
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            date_str = request.GET.get("date", "")
            try:
                session_date = date.fromisoformat(date_str)
            except ValueError:
                return JsonResponse({"error": "invalid date"}, status=400)
            return JsonResponse({"tables": _therapies_for_date(session_date)})

        return render(request, self.template_name, self._base_context(request))

    def post(self, request):
        selected_date_str = request.POST.get("session_date", "")
        try:
            session_date = date.fromisoformat(selected_date_str)
        except ValueError:
            messages.error(request, "Dată invalidă.")
            return render(request, self.template_name, self._base_context(request))

        import re as _re
        # Parse tables — new structure: tables[N][therapist], tables[N][slots][M][child|start_time]
        raw = request.POST
        tables = {}
        for key, value in raw.items():
            m = _re.match(r'tables\[(\d+)\]\[therapist\]$', key)
            if m:
                t_idx = m.group(1)
                tables.setdefault(t_idx, {"therapist": None, "slots": {}})
                tables[t_idx]["therapist"] = value
                continue
            m = _re.match(r'tables\[(\d+)\]\[slots\]\[(\d+)\]\[(child|start_time)\]$', key)
            if m:
                t_idx, s_idx, field = m.group(1), m.group(2), m.group(3)
                tables.setdefault(t_idx, {"therapist": None, "slots": {}})
                tables[t_idx]["slots"].setdefault(s_idx, {})
                tables[t_idx]["slots"][s_idx][field] = value

        errors = []
        to_create = []
        table_entries_by_idx = {}  # Track entries per table index
        therapist_by_idx = {}  # Track therapist objects for each table index
        incomplete_tables_dict = {}  # Store incomplete table data indexed by t_idx
        processed_indices = set()  # Track which table indices have been processed
        
        # Create index mapping: sort table indices to get correct display order
        sorted_table_indices = sorted(tables.keys(), key=lambda x: int(x))
        idx_to_display_num = {idx: i + 1 for i, idx in enumerate(sorted_table_indices)}

        # First pass: validate tables and build therapist map (in sorted order)
        for t_idx in sorted_table_indices:
            table = tables[t_idx]
            therapist_id = table.get("therapist")
            if not therapist_id:
                # Check if the table has any slots with data
                has_slots = any(slot.get("child") or slot.get("start_time") for slot in table["slots"].values())
                if has_slots:
                    errors.append(f"Tabela {idx_to_display_num[t_idx]}: terapeut neselectat.")
                    # Save incomplete table for restoration
                    slots_list = []
                    for s_idx, slot in sorted(table["slots"].items(), key=lambda x: int(x[0])):
                        child_id = slot.get("child", "")
                        start_time_str = slot.get("start_time", "")
                        if child_id or start_time_str:
                            slots_list.append({"child_id": child_id, "start_time": start_time_str})
                    incomplete_tables_dict[t_idx] = {"therapist_id": None, "slots": slots_list}
                    processed_indices.add(t_idx)  # Mark as processed
                continue
            try:
                therapist = Therapist.objects.get(pk=therapist_id)
                therapist_by_idx[t_idx] = therapist
            except Therapist.DoesNotExist:
                errors.append(f"Tabela {idx_to_display_num[t_idx]}: terapeut invalid.")
                processed_indices.add(t_idx)  # Mark as processed
                continue

        # Second pass: validate slots and collect entries (in sorted order)
        for t_idx in sorted_table_indices:
            # Skip tables already processed in first pass
            if t_idx in processed_indices:
                continue
            
            table = tables[t_idx]
            if t_idx not in therapist_by_idx:
                # Table has no therapist - only save if it has slot data (to show errors)
                slots_list = []
                for s_idx, slot in sorted(table["slots"].items(), key=lambda x: int(x[0])):
                    child_id = slot.get("child", "")
                    start_time_str = slot.get("start_time", "")
                    if child_id or start_time_str:
                        slots_list.append({"child_id": child_id, "start_time": start_time_str})
                
                # Only save table if it has slot data (not completely empty)
                if slots_list:
                    incomplete_tables_dict[t_idx] = {
                        "therapist_id": table.get("therapist"),
                        "slots": slots_list
                    }
                continue

            therapist = therapist_by_idx[t_idx]

            table_entries = []
            table_has_errors = False
            
            for s_idx, slot in sorted(table["slots"].items(), key=lambda x: int(x[0])):
                child_id = slot.get("child", "")
                start_time_str = slot.get("start_time", "")

                # Check for missing child
                if not child_id and not start_time_str:
                    # Both empty, skip this slot silently
                    continue
                
                # Calculate end time for interval display
                end_time_str = ""
                if start_time_str:
                    try:
                        h, m = map(int, start_time_str.split(":"))
                        end_h = (h + 2) % 24
                        end_time_str = f"{end_h:02d}:00"
                        interval_str = f"{start_time_str}-{end_time_str}"
                    except (ValueError, AttributeError):
                        interval_str = start_time_str
                else:
                    interval_str = "necunoscut"

                if not child_id:
                    errors.append(f"Tabela {idx_to_display_num[t_idx]}, interval {interval_str}: pacient neselectat.")
                    table_has_errors = True
                    continue
                if not start_time_str:
                    errors.append(f"Tabela {idx_to_display_num[t_idx]}: oră lipsă pentru pacientul selectat.")
                    table_has_errors = True
                    continue

                try:
                    h, m_val = map(int, start_time_str.split(":"))
                    start_t = time(h, m_val)
                except (ValueError, AttributeError):
                    errors.append(f"Tabela {idx_to_display_num[t_idx]}, interval {interval_str}: oră invalidă.")
                    table_has_errors = True
                    continue
                try:
                    child = Child.objects.get(pk=child_id, status="activ")
                except Child.DoesNotExist:
                    errors.append(f"Tabela {idx_to_display_num[t_idx]}, interval {interval_str}: pacient invalid sau arhivat.")
                    table_has_errors = True
                    continue

                table_entries.append(Therapy(
                    child=child,
                    therapist=therapist,
                    date=session_date,
                    start_time=start_t,
                ))

            # If this table has errors, save it for restoration
            if table_has_errors:
                slots_list = []
                for s_idx, slot in sorted(table["slots"].items(), key=lambda x: int(x[0])):
                    child_id = slot.get("child", "")
                    start_time_str = slot.get("start_time", "")
                    if child_id or start_time_str:
                        slots_list.append({"child_id": child_id, "start_time": start_time_str})
                incomplete_tables_dict[t_idx] = {
                    "therapist_id": therapist.id,
                    "slots": slots_list
                }
            else:
                # Only add entries if no errors for this table
                table_entries_by_idx[t_idx] = (therapist, table_entries)
                to_create.extend(table_entries)

        # Third pass: check for overlapping sessions for the same therapist
        from datetime import timedelta as td
        therapist_slots = {}  # { therapist_id: [list of (start_time, end_time, therapy_obj)] }
        for therapy in to_create:
            t_id = therapy.therapist_id
            if t_id not in therapist_slots:
                therapist_slots[t_id] = []
            end_time = time((therapy.start_time.hour + 2) % 24, therapy.start_time.minute)
            therapist_slots[t_id].append((therapy.start_time, end_time, therapy))

        # Check for overlaps
        conflict_table_indices = set()  # Track which table indices have conflicts
        for t_id, slots in therapist_slots.items():
            for i, (start1, end1, therapy1) in enumerate(slots):
                for start2, end2, therapy2 in slots[i+1:]:
                    # Two sessions overlap if: start1 < end2 AND start2 < end1
                    if start1 < end2 and start2 < end1:
                        interval1 = f"{start1.strftime('%H:%M')}-{end1.strftime('%H:%M')}"
                        interval2 = f"{start2.strftime('%H:%M')}-{end2.strftime('%H:%M')}"
                        errors.append(
                            f"Conflict de orar pentru {therapy1.therapist}: "
                            f"intervalele {interval1} și {interval2} se suprapun."
                        )
                        # Mark all therapist's tables as conflicting
                        for t_idx, (therapist_obj, entries) in table_entries_by_idx.items():
                            if therapist_obj.id == t_id:
                                conflict_table_indices.add(t_idx)

        # If there are conflicts, save conflicting tables and clear to_create
        if conflict_table_indices:
            for t_idx in sorted(conflict_table_indices, key=lambda x: int(x)):
                therapist_obj, table_entries = table_entries_by_idx[t_idx]
                slots_list = []
                for s_idx, slot in sorted(tables[t_idx]["slots"].items(), key=lambda x: int(x[0])):
                    child_id = slot.get("child", "")
                    start_time_str = slot.get("start_time", "")
                    if child_id or start_time_str:
                        slots_list.append({"child_id": child_id, "start_time": start_time_str})
                incomplete_tables_dict[t_idx] = {
                    "therapist_id": therapist_obj.id,
                    "slots": slots_list
                }
            to_create = []  # Clear entries with conflicts

        if errors:
            # Save good entries to DB
            if to_create:
                with transaction.atomic():
                    Therapy.objects.filter(date=session_date).delete()
                    Therapy.objects.bulk_create(to_create)
                    messages.success(
                        request,
                        f"{len(to_create)} ședință(e) au fost salvate. "
                        f"Tabelele cu erori rămân pentru remediere."
                    )
            
            # Show error messages
            for e in errors:
                messages.error(request, e)
            
            # Convert incomplete_tables_dict to sorted list
            incomplete_tables = [incomplete_tables_dict[t_idx] for t_idx in sorted_table_indices if t_idx in incomplete_tables_dict]
            
            ctx = self._base_context(request)
            ctx['incomplete_tables'] = json.dumps(incomplete_tables)
            ctx['session_date'] = selected_date_str
            return render(request, self.template_name, ctx)

        with transaction.atomic():
            # Replace: delete all sessions for this date, then bulk-insert the new ones
            deleted, _ = Therapy.objects.filter(date=session_date).delete()
            Therapy.objects.bulk_create(to_create)

        messages.success(
            request,
            f"Orar salvat pentru {session_date}: {len(to_create)} ședință(e) "
            f"({deleted} înlocuite)." if deleted else
            f"Orar salvat pentru {session_date}: {len(to_create)} ședință(e)."
        )
        return redirect("admin:therapy_batch")


# ──────────────────────────────────────────────────────────────────────────────
# Daily report
# ──────────────────────────────────────────────────────────────────────────────
@method_decorator(staff_member_required, name="dispatch")
class RaportZilnicView(View):
    template_name = "admin/core/raport_zilnic.html"

    def get(self, request):
        today = date.today()
        selected = request.GET.get("data", today.isoformat())
        try:
            selected_date = date.fromisoformat(selected)
        except ValueError:
            selected_date = today

        therapies = (
            Therapy.objects
            .filter(date=selected_date)
            .select_related("child")
        )

        counts = {}
        names  = {}
        for t in therapies:
            cid = t.child_id
            counts[cid] = counts.get(cid, 0) + 1
            names[cid]  = f"{t.child.last_name} {t.child.first_name}"

        rows = [
            {"child": names[cid], "sessions": cnt, "hours": cnt * 2}
            for cid, cnt in sorted(counts.items(), key=lambda x: names[x[0]].lower())
        ]

        ctx = _report_context(request, f"Raport zilnic – {selected_date.strftime('%d.%m.%Y')}")
        ctx.update({
            "selected_date": selected_date,
            "today": today,
            "rows": rows,
            "total_hours": sum(r["hours"] for r in rows),
        })
        return render(request, self.template_name, ctx)


def admin_site_context(request):
    from django.contrib import admin as _admin
    return {
        "has_permission": request.user.is_active and request.user.is_staff,
        "site_header": _admin.site.site_header,
        "site_title": _admin.site.site_title,
        "title": "Editează/Adaugă orar",
        "opts": {"app_label": "core"},
    }


# ─────────────────────────────────────────────────────────────────────────────
# Report helpers
# ─────────────────────────────────────────────────────────────────────────────

MONTHS_RO = [
    "", "Ianuarie", "Februarie", "Martie", "Aprilie", "Mai", "Iunie",
    "Iulie", "August", "Septembrie", "Octombrie", "Noiembrie", "Decembrie",
]


def _available_years():
    from django.db.models import Min, Max
    agg = Therapy.objects.aggregate(mn=Min("date__year"), mx=Max("date__year"))
    mn = agg["mn"] or date.today().year
    mx = agg["mx"] or date.today().year
    return list(range(mn, mx + 1))


def _report_context(request, title):
    from django.contrib import admin as _admin
    return {
        "has_permission": request.user.is_active and request.user.is_staff,
        "site_header": _admin.site.site_header,
        "site_title": _admin.site.site_title,
        "title": title,
        "opts": {"app_label": "core"},
    }


@method_decorator(staff_member_required, name="dispatch")
class RaportLunarView(View):
    template_name = "admin/core/raport_lunar.html"

    def get(self, request):
        today = date.today()
        year  = int(request.GET.get("an",  today.year))
        month = int(request.GET.get("luna", today.month))

        _, days_in_month = calendar.monthrange(year, month)
        days = list(range(1, days_in_month + 1))

        # Fetch all therapies for the month
        therapies = (
            Therapy.objects
            .filter(date__year=year, date__month=month)
            .select_related("child", "therapist")
        )

        # Build pivot: {(child_id, therapist_id): {day: count}}
        pivot = {}
        meta  = {}  # key -> (child_display, therapist_display, cnp)
        for t in therapies:
            key = (t.child_id, t.therapist_id)
            if key not in pivot:
                pivot[key] = {}
                meta[key] = (
                    f"{t.child.last_name} {t.child.first_name}",
                    f"{t.therapist.last_name} {t.therapist.first_name}",
                    t.child.cnp,
                )
            d = t.date.day
            pivot[key][d] = pivot[key].get(d, 0) + 1

        filter_child     = request.GET.get("child", "").strip()
        filter_therapist = request.GET.get("therapist", "").strip()
        children_list    = sorted({v[0] for v in meta.values()}, key=str.lower)
        therapists_list  = sorted({v[1] for v in meta.values()}, key=str.lower)

        rows = []
        for key in sorted(meta, key=lambda k: (meta[k][0].lower(), meta[k][1].lower())):
            child_name, therapist_name, cnp = meta[key]
            if filter_child and child_name != filter_child:
                continue
            if filter_therapist and therapist_name != filter_therapist:
                continue
            counts = [(pivot[key].get(d, 0) * 2) or "" for d in days]
            total  = sum(v for v in counts if v)
            rows.append({
                "child":     child_name,
                "therapist": therapist_name,
                "cnp":       cnp,
                "counts":    counts,
                "total":     total,
            })

        ctx = _report_context(request, f"Raport lunar – {MONTHS_RO[month]} {year}")
        ctx.update({
            "year": year, "month": month,
            "month_name": MONTHS_RO[month],
            "days": days,
            "rows": rows,
            "years": _available_years(),
            "months": list(enumerate(MONTHS_RO))[1:],
            "filter_child": filter_child,
            "filter_therapist": filter_therapist,
            "children_list": children_list,
            "therapists_list": therapists_list,
        })
        return render(request, self.template_name, ctx)


@method_decorator(staff_member_required, name="dispatch")
class RaportAnualView(View):
    template_name = "admin/core/raport_anual.html"

    def get(self, request):
        today = date.today()
        year  = int(request.GET.get("an", today.year))

        therapies = (
            Therapy.objects
            .filter(date__year=year)
            .select_related("child", "therapist")
        )

        pivot = {}
        meta  = {}
        for t in therapies:
            key = (t.child_id, t.therapist_id)
            if key not in pivot:
                pivot[key] = {}
                meta[key] = (
                    f"{t.child.last_name} {t.child.first_name}",
                    f"{t.therapist.last_name} {t.therapist.first_name}",
                    t.child.cnp,
                )
            m = t.date.month
            pivot[key][m] = pivot[key].get(m, 0) + 1
        filter_child     = request.GET.get("child", "").strip()
        filter_therapist = request.GET.get("therapist", "").strip()
        children_list    = sorted({v[0] for v in meta.values()}, key=str.lower)
        therapists_list  = sorted({v[1] for v in meta.values()}, key=str.lower)

        rows = []
        for key in sorted(meta, key=lambda k: (meta[k][0].lower(), meta[k][1].lower())):
            child_name, therapist_name, cnp = meta[key]
            if filter_child and child_name != filter_child:
                continue
            if filter_therapist and therapist_name != filter_therapist:
                continue
            counts = [(pivot[key].get(m, 0) * 2) or "" for m in range(1, 13)]
            total  = sum(v for v in counts if v)
            rows.append({
                "child":     child_name,
                "therapist": therapist_name,
                "cnp":       cnp,
                "counts":    counts,
                "total":     total,
            })

        ctx = _report_context(request, f"Raport anual – {year}")
        ctx.update({
            "year": year,
            "month_names": MONTHS_RO[1:],
            "rows": rows,
            "years": _available_years(),
            "months": list(enumerate(MONTHS_RO))[1:],
            "filter_child": filter_child,
            "filter_therapist": filter_therapist,
            "children_list": children_list,
            "therapists_list": therapists_list,
        })
        return render(request, self.template_name, ctx)


# ──────────────────────────────────────────────────────────────────────────────
# Weekly report
# ──────────────────────────────────────────────────────────────────────────────
DAYS_RO = ["Luni", "Marți", "Miercuri", "Joi", "Vineri", "Sâmbătă", "Duminică"]


def _week_bounds(year, week):
    """Return (monday, sunday) for ISO week."""
    monday = date.fromisocalendar(year, week, 1)
    sunday = monday + timedelta(days=6)
    return monday, sunday


def _weeks_for_year(year):
    """Return list of ISO week numbers (1-52 or 1-53) for the given year."""
    # ISO week 53 exists only if Dec 28 falls in week 53
    last_week = date(year, 12, 28).isocalendar().week
    return list(range(1, last_week + 1))


@method_decorator(staff_member_required, name="dispatch")
class RaportSaptamanaiView(View):
    template_name = "admin/core/raport_saptamanal.html"

    def get(self, request):
        today = date.today()
        iso = today.isocalendar()
        year = int(request.GET.get("an", iso.year))
        week = int(request.GET.get("sapt", iso.week))

        monday, sunday = _week_bounds(year, week)
        days = [monday + timedelta(days=i) for i in range(7)]

        therapies = (
            Therapy.objects
            .filter(date__gte=monday, date__lte=sunday)
            .select_related("child", "therapist")
        )

        pivot = {}
        meta  = {}
        for t in therapies:
            key = (t.child_id, t.therapist_id)
            if key not in pivot:
                pivot[key] = {}
                meta[key] = (
                    f"{t.child.last_name} {t.child.first_name}",
                    f"{t.therapist.last_name} {t.therapist.first_name}",
                    t.child.cnp,
                )
            pivot[key][t.date] = pivot[key].get(t.date, 0) + 1

        filter_child     = request.GET.get("child", "").strip()
        filter_therapist = request.GET.get("therapist", "").strip()
        children_list    = sorted({v[0] for v in meta.values()}, key=str.lower)
        therapists_list  = sorted({v[1] for v in meta.values()}, key=str.lower)

        rows = []
        for key in sorted(meta, key=lambda k: (meta[k][0].lower(), meta[k][1].lower())):
            child_name, therapist_name, cnp = meta[key]
            if filter_child and child_name != filter_child:
                continue
            if filter_therapist and therapist_name != filter_therapist:
                continue
            counts = [(pivot[key].get(d, 0) * 2) or "" for d in days]
            total  = sum(v for v in counts if v)
            rows.append({
                "child":     child_name,
                "therapist": therapist_name,
                "cnp":       cnp,
                "counts":    counts,
                "total":     total,
            })

        weeks = _weeks_for_year(year)
        ctx = _report_context(request, f"Raport săptămânal – S{week} {year}")
        ctx.update({
            "year": year, "week": week,
            "monday": monday, "sunday": sunday,
            "days": days,
            "days_ro": DAYS_RO,
            "rows": rows,
            "years": _available_years(),
            "weeks": weeks,
            "filter_child": filter_child,
            "filter_therapist": filter_therapist,
            "children_list": children_list,
            "therapists_list": therapists_list,
        })
        return render(request, self.template_name, ctx)


# ──────────────────────────────────────────────────────────────────────────────
# Excel export helpers
# ──────────────────────────────────────────────────────────────────────────────
def _make_excel(title, headers, rows):
    """Build an openpyxl workbook and return bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="318CE7")
    total_fill   = PatternFill("solid", fgColor="D0E8FB")
    thin         = Side(style="thin", color="CCCCCC")
    cell_border  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.border = cell_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for r_idx, row in enumerate(rows, 2):
        is_total_col = False
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val if val != "" else None)
            cell.border = cell_border
            # Last column = total
            if c_idx == len(row):
                cell.fill = total_fill
                cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center" if c_idx > 3 else "left")
        # Alternating row shading
        if r_idx % 2 == 0:
            for c_idx in range(1, len(row)):
                ws.cell(row=r_idx, column=c_idx).fill = PatternFill("solid", fgColor="F5F8FF")

    # Auto-fit column widths (approx)
    for col_cells in ws.columns:
        length = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 8), 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _excel_response(data: bytes, filename: str) -> HttpResponse:
    resp = HttpResponse(
        data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@staff_member_required
def raport_zilnic_excel(request):
    today = date.today()
    selected = request.GET.get("data", today.isoformat())
    try:
        selected_date = date.fromisoformat(selected)
    except ValueError:
        selected_date = today
    hours_mode = request.GET.get("ore", "0") == "1"

    therapies = Therapy.objects.filter(date=selected_date).select_related("child")
    counts, names = {}, {}
    for t in therapies:
        cid = t.child_id
        counts[cid] = counts.get(cid, 0) + 1
        names[cid] = f"{t.child.last_name} {t.child.first_name}"

    label = "Ore terapie" if hours_mode else "Şedințe terapie"
    headers = ["Pacient", label]
    rows = [
        [names[cid], (cnt * 2 if hours_mode else cnt)]
        for cid, cnt in sorted(counts.items(), key=lambda x: names[x[0]].lower())
    ]
    suffix = "_ore" if hours_mode else "_sedinte"
    ts = datetime.now().strftime("%d_%m_%Y_%H_%M")
    data = _make_excel(f"Zilnic {selected_date}", headers, rows)
    return _excel_response(data, f"raport_zilnic_{selected_date}{suffix}_{ts}.xlsx")


@staff_member_required
def raport_lunar_excel(request):
    today = date.today()
    year  = int(request.GET.get("an",  today.year))
    month = int(request.GET.get("luna", today.month))
    hours_mode = request.GET.get("ore", "0") == "1"

    _, days_in_month = calendar.monthrange(year, month)
    days = list(range(1, days_in_month + 1))

    therapies = (
        Therapy.objects
        .filter(date__year=year, date__month=month)
        .select_related("child", "therapist")
    )
    pivot, meta = {}, {}
    for t in therapies:
        key = (t.child_id, t.therapist_id)
        if key not in pivot:
            pivot[key] = {}
            meta[key] = (
                f"{t.child.last_name} {t.child.first_name}",
                f"{t.therapist.last_name} {t.therapist.first_name}",
                t.child.cnp,
            )
        d = t.date.day
        pivot[key][d] = pivot[key].get(d, 0) + 1

    filter_child     = request.GET.get("child", "").strip()
    filter_therapist = request.GET.get("therapist", "").strip()
    mul = 2 if hours_mode else 1
    label = "Ore" if hours_mode else "Şedințe"
    headers = ["Pacient", "Terapeut", "CNP"] + [str(d) for d in days] + [f"Total {label}"]
    rows = []
    for key in sorted(meta, key=lambda k: (meta[k][0].lower(), meta[k][1].lower())):
        child_name, therapist_name, cnp = meta[key]
        if filter_child and child_name != filter_child:
            continue
        if filter_therapist and therapist_name != filter_therapist:
            continue
        counts = [pivot[key].get(d, "") for d in days]
        total  = sum(v for v in counts if v) * mul
        rows.append([child_name, therapist_name, cnp] + [(v * mul if v else "") for v in counts] + [total])

    suffix = "_ore" if hours_mode else "_sedinte"
    ts = datetime.now().strftime("%d_%m_%Y_%H_%M")
    data = _make_excel(f"{MONTHS_RO[month]} {year}", headers, rows)
    return _excel_response(data, f"raport_lunar_{year}_{month:02d}{suffix}_{ts}.xlsx")


@staff_member_required
def raport_anual_excel(request):
    today = date.today()
    year  = int(request.GET.get("an", today.year))
    hours_mode = request.GET.get("ore", "0") == "1"

    therapies = (
        Therapy.objects
        .filter(date__year=year)
        .select_related("child", "therapist")
    )
    pivot, meta = {}, {}
    for t in therapies:
        key = (t.child_id, t.therapist_id)
        if key not in pivot:
            pivot[key] = {}
            meta[key] = (
                f"{t.child.last_name} {t.child.first_name}",
                f"{t.therapist.last_name} {t.therapist.first_name}",
                t.child.cnp,
            )
        m = t.date.month
        pivot[key][m] = pivot[key].get(m, 0) + 1

    filter_child     = request.GET.get("child", "").strip()
    filter_therapist = request.GET.get("therapist", "").strip()
    mul = 2 if hours_mode else 1
    label = "Ore" if hours_mode else "Şedințe"
    headers = ["Pacient", "Terapeut", "CNP"] + MONTHS_RO[1:] + [f"Total {label}"]
    rows = []
    for key in sorted(meta, key=lambda k: (meta[k][0].lower(), meta[k][1].lower())):
        child_name, therapist_name, cnp = meta[key]
        if filter_child and child_name != filter_child:
            continue
        if filter_therapist and therapist_name != filter_therapist:
            continue
        counts = [pivot[key].get(m, "") for m in range(1, 13)]
        total  = sum(v for v in counts if v) * mul
        rows.append([child_name, therapist_name, cnp] + [(v * mul if v else "") for v in counts] + [total])

    suffix = "_ore" if hours_mode else "_sedinte"
    ts = datetime.now().strftime("%d_%m_%Y_%H_%M")
    data = _make_excel(f"Anual {year}", headers, rows)
    return _excel_response(data, f"raport_anual_{year}{suffix}_{ts}.xlsx")


@staff_member_required
def raport_saptamanal_excel(request):
    today = date.today()
    iso = today.isocalendar()
    year = int(request.GET.get("an", iso.year))
    week = int(request.GET.get("sapt", iso.week))
    hours_mode = request.GET.get("ore", "0") == "1"

    monday, sunday = _week_bounds(year, week)
    days = [monday + timedelta(days=i) for i in range(7)]

    therapies = (
        Therapy.objects
        .filter(date__gte=monday, date__lte=sunday)
        .select_related("child", "therapist")
    )
    pivot, meta = {}, {}
    for t in therapies:
        key = (t.child_id, t.therapist_id)
        if key not in pivot:
            pivot[key] = {}
            meta[key] = (
                f"{t.child.last_name} {t.child.first_name}",
                f"{t.therapist.last_name} {t.therapist.first_name}",
                t.child.cnp,
            )
        pivot[key][t.date] = pivot[key].get(t.date, 0) + 1

    mul = 2 if hours_mode else 1
    label = "Ore" if hours_mode else "Şedințe"
    filter_child     = request.GET.get("child", "").strip()
    filter_therapist = request.GET.get("therapist", "").strip()
    day_labels = [f"{DAYS_RO[d.weekday()]} {d.strftime('%d.%m')}" for d in days]
    headers = ["Pacient", "Terapeut", "CNP"] + day_labels + [f"Total {label}"]
    rows = []
    for key in sorted(meta, key=lambda k: (meta[k][0].lower(), meta[k][1].lower())):
        child_name, therapist_name, cnp = meta[key]
        if filter_child and child_name != filter_child:
            continue
        if filter_therapist and therapist_name != filter_therapist:
            continue
        counts = [pivot[key].get(d, "") for d in days]
        total  = sum(v for v in counts if v) * mul
        rows.append([child_name, therapist_name, cnp] + [(v * mul if v else "") for v in counts] + [total])

    suffix = "_ore" if hours_mode else "_sedinte"
    ts = datetime.now().strftime("%d_%m_%Y_%H_%M")
    data = _make_excel(f"S{week} {year}", headers, rows)
    return _excel_response(data, f"raport_saptamanal_{year}_S{week:02d}{suffix}_{ts}.xlsx")


# ──────────────────────────────────────────────────────────────────────────────
# PDF export for therapy schedule
# ──────────────────────────────────────────────────────────────────────────────
@staff_member_required
def therapy_schedule_pdf(request):
    """Generate PDF with therapy schedule for a given date."""
    selected_date_str = request.GET.get("date", "")
    try:
        selected_date = date.fromisoformat(selected_date_str)
    except (ValueError, TypeError):
        messages.error(request, "Dată invalidă.")
        return redirect("admin:core_therapy_changelist")
    
    from core.pdf import generate_therapy_pdf
    
    pdf_data = generate_therapy_pdf(selected_date)
    
    response = HttpResponse(pdf_data, content_type='application/pdf')
    date_str = selected_date.strftime('%d_%m_%Y')
    response['Content-Disposition'] = f'attachment; filename="orar_terapii_{date_str}.pdf"'
    return response
